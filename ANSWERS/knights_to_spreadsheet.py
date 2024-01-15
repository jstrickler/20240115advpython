import openpyxl as px

COLUMN_HEADINGS = 'Name', 'Title', 'Favorite Color', 'Quest', 'Comment'
NAME_FONT = px.styles.Font(color='FF0000')   # name is red
COMMENT_FONT = px.styles.Font(italic=True)   # comment is italics

wb = px.Workbook()  # create new workbook

ws = wb.active  # select active (and only) worksheet

ws.title = 'knights'  # set worksheet title

for column, heading in enumerate(COLUMN_HEADINGS, 1):
    ws.cell(row=1, column=column).value = heading
    ws.cell(row=1, column=column).font = px.styles.Font(bold=True)

with open('../DATA/knights.txt') as knights_in:
    for line in knights_in:
        fields = line.rstrip().split(':')
        ws.append(fields)
        ws.cell(ws.max_row, 1).font = NAME_FONT
        ws.cell(ws.max_row, 5).font = COMMENT_FONT

    # the "hard way":
    # for row, line in enumerate(knights_in, 2):
        # name, title, color, quest, comment = line[:-1].split(':')
        # ws.cell(row, 1).value = name
        # ws.cell(row, 1).font = px.styles.Font(color='FFFF0000')
        # ws.cell(row, 2).value = title
        # ws.cell(row, 3).value = color
        # ws.cell(row, 4).value = quest
        # ws.cell(row, 5).value = comment
        # ws.cell(row, 5).font = px.styles.Font(italic=True)

wb.save('knights.xlsx')


